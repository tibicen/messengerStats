# -*- coding: utf-8 -*-
"""
Created on Thu Jul 21 15:08:39 2016

@author: tibicen
"""
import time
import string
import os

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

print('imported.')


def splitFile():
    '''splits 'messages.htm' into smallet files easier to open and inspect.
       Only for inspecting, not a part of the rest.
    '''
    with open('messages.htm', 'r', encoding='UTF-8') as f:
        filenr = 0
        nr = 0
        tmp = ''
        for line in f:
            tmp += line
            nr += 1
            if nr == 130:
                partfile = open('messages%d.htm' % filenr, 'w',
                                encoding='UTF-8')
                partfile.write(tmp)
                partfile.close()
                filenr += 1
                tmp = ''
                nr = 0
        partfile = open('messages%d.htm' % filenr, 'w', encoding='UTF-8')
        partfile.write(tmp)
        partfile.close()
        filenr += 1
        tmp = ''
        nr = ''


def splitIntoTxtFiles():
    '''splits 'messages.htm' into txt files for every person.
    in each file are only words
    '''
    # %% OPENING
    text = ''
    print('opening file...', end='\t')
    t1 = time.time()
    with open('messages.htm', 'r', encoding='UTF-8') as f:
        for line in f:
            text += line
    t2 = time.time()
    print('file opened in %.2f.' % (t2 - t1))
    # %% CREATING SOUP
    print("creating soup...", end='\t')
    soup = BeautifulSoup(text, 'lxml')
    t3 = time.time()
    del text
    print('text souped in %.2f.' % (t3 - t2))
    # %% CREATING DICT
    print('creating dict...', end='\t')
    usersDict = {}
    threads = soup.findAll('div', {'class': 'thread'})
    jobCount = len(threads)
    if 'realPersons' not in os.listdir():
        os.mkdir('realPersons')
    if 'randomConversations' not in os.listdir():
        os.mkdir('randomConversations')
    for thread in threads:
        #        print('%d threads left.' % jobCount)
        jobCount -= 1
        childs = list(thread.recursiveChildGenerator())
        maxIter = len(childs)
        for n, child in enumerate(childs):
            #            print(child)
            #            input()
            if n + 7 + 1 > maxIter:
                break
            try:
                if child.has_attr('class'):
                    if 'user' in child.attrs['class']:
                        if usersDict.get(childs[n + 1]):
                            usersDict[childs[n + 1]] += [childs[n + 5] + '\n']
                        else:
                            usersDict[childs[n + 1]] = [childs[n + 5] + '\n']
            except(AttributeError, TypeError):
                pass
    del threads
    t4 = time.time()
    print('dict created in %.2f.' % (t4 - t3))
    # %% SAVING
    print('saving...', end='\t\t')
    for user, text in usersDict.items():
        if user.startswith(tuple([x for x in string.digits])):
            folder = 'randomConversations'
        else:
            folder = 'realPersons'
        f = open(os.path.join(folder, user + '.txt'), 'w', encoding='UTF-8')
        f.writelines(text)
        f.close()
    t5 = time.time()
    del usersDict
    print('saved in %.2f.' % (t5 - t4))


# %%
def getWordsDict(ppl, filename):
    ''' populates ppl dict with records form filename'''
    f = open(os.path.join('realPersons', filename), 'r', encoding='UTF-8')
    personName = filename.rstrip('.txt')
    ppl[personName] = {}
    text = f.read().lower()
    nonwords = string.digits + string.punctuation + string.whitespace
    for s in nonwords:
        text = text.replace(s, ' ')
    wordList = []
    # deleting useles words with no meaning
    # PUT YOUR OWN WORDS HERE
    badWords = [x for x in string.ascii_lowercase] + \
        ['', 'ma', 'w', 'i', 'na', 'z', 'a', 'bo', 'o', 'za', 'ze',
         'od', 'po', 'na', 'pod', 'no', 'do', 'co', 'że', 'jak',
         'czy', 'sie', 'już', 'to', 'się', 'też', 'coś', 'żeby',
         'są', 'we', 'te', 'ale', 'więc', 'tym', 'tam', 'com', 'http',
         'https', 'www', 'dla', 'pl', 'at', 'and', 'of', 'in', 'for', 'so',
         'am', 'so']
    for w in text.split(' '):
        if w not in badWords:
            wordList.append(w)
    for w in wordList:
        if ppl[personName].get(w):
            ppl[personName][w] += 1
        else:
            ppl[personName][w] = 1
        if ppl['all'].get(w):
            ppl['all'][w] += 1
        else:
            ppl['all'][w] = 1
    del wordList


# %%
def createXLS(ppl):
    '''sorts  all data and saves it into xls file'''
    wb = Workbook()
    sh = wb.active
    # type of sorting here
    wordDict = sorted([(k,
                        sorted(v.items(), key=lambda x: x[1], reverse=True),
                        len(v.items())) for (k, v) in ppl.items()],
                      key=lambda x: x[2], reverse=True)

    col = 1
    # CUSTOMIZE ALL THE WORDS FOR YOUR CONVINIENCE
    for name, words, count in wordDict:
        sh.cell(row=1, column=col).value = name
        for n, word in enumerate(words):
            cell = sh.cell(row=2 + n, column=col)
            cell.value = word[0]
            if word[0] == 'nie':
                cell.fill = PatternFill(
                    fill_type='solid', start_color='fc5834')
                cell.font = Font(color='000000')
            elif word[0] == 'tak':
                cell.fill = PatternFill(
                    fill_type='solid', start_color='a1c870')
                cell.font = Font(color='000000')
            elif word[0] in ('ja', 'mi', 'mam', 'mnie', 'jestem', 'wiem',
                             'bede', 'będę', 'mogę', 'chce', 'mialam',
                             'bym', 'sam', 'mój', 'chcę', 'moje', 'mną',
                             'chciałem', 'my', 'i'):
                cell.fill = PatternFill(
                    fill_type='solid', start_color='7ee0e9')
                cell.font = Font(color='ffffff')
            elif word[0] in ('ci', 'ty', 'masz', 'ciebie', 'jesteś',
                             'możesz', 'chcesz', 'cię', 'będziesz', 'you'):
                cell.fill = PatternFill(
                    fill_type='solid', start_color='d57ee3')
                cell.font = Font(color='ffffff')
            elif word[0] in ('mu', 'jej', 'on', 'ten', 'go', 'ona'):
                cell.fill = PatternFill(
                    fill_type='solid', start_color='7b89e8')
                cell.font = Font(color='ffffff')
            else:
                cell.font = Font(color='5c5c5c')
        col += 1
    wb.save('ppl.xls')


if __name__ == '__main__':
    # splitIntoTxtFiles()
    ppl = {}  # ppl['person name'] = {'word1': wordCountNr}
    ppl['all'] = {}
    files = os.listdir('realPersons')
    # populate ppl dict with people records and words
    t1 = time.time()
    print('populating ppl dict...', end='\t')
    for filename in files:
        getWordsDict(ppl, filename)
    t2 = time.time()
    print('done in %.2f.' % (t2 - t1))
    print('creating xls file...', end='\t')
    createXLS(ppl)
    t3 = time.time()
    print('done and saved in %.2f.' % (t3 - t2))
